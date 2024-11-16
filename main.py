from pypdf import PdfReader
import re
from datetime import datetime

from strings_to_remove import strings_to_remove
from patterns_to_remove import patterns_to_remove

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, NamedStyle, PatternFill

# Set parameters

date = "01092024"  # Change to match date in receipt file name
person1 = "Alice"
person2 = "Bob"

# Extract text from PDF

pdf_name = "receipt-" + date + ".pdf"

reader = PdfReader(pdf_name)
text = ""
for page in reader.pages:
    text += page.extract_text() + "\n"

# Remove unused text

for string in strings_to_remove:
    text = text.replace(string, " ")

# Remove unused patterns

for pattern in patterns_to_remove:
    text = re.sub(pattern, "", text)

# Extract delivery charge and remove it from text

delivery_charge = re.findall(r"Picking, packing and delivery £\d+.\d+", text)[0]
text = text.replace(delivery_charge, "")
delivery_charge = float(delivery_charge.replace("Picking, packing and delivery £", ""))

# Extract total charge and remove it from text

total_charge = re.findall(r"Total charge £\d+.\d+", text)[0]
text = text.replace(total_charge, "")
total_charge = float(total_charge.replace("Total charge £", ""))

# Remove some more unused text

text = text.replace("to pay", "")

# Split by item purchased

text = re.split(r"(\d+\.\d+ )", text)

# Create list of 2-element lists

list_of_lists = []

for element in text:
    if text.index(element) % 2 == 0:
        my_list = [element.strip()]
    else:
        my_list.append(float(element))
        list_of_lists.append(my_list)
        my_list = []

# Add delivery charge

list_of_lists.append(["DELIVERY CHARGE", delivery_charge])

# Calculate total spending and check it against total charge

total_spending = 0

for item in list_of_lists:
    total_spending += item[1]

assert round(total_charge, 2) == round(total_spending, 2)

# Create Excel file

wb = Workbook()
ws = wb.active

# Add data to Excel

for row in list_of_lists:
    ws.append(row)

# Adjust column widths

ws.column_dimensions['A'].width = 70

# Add formulas

ws['E1'] = person1
ws['E2'] = person2
ws['F1'] = f'=SUMIF(C1:C100,"={person1}",B1:B100)+SUMIF(C1:C100,"=Both",B1:B100)/2'
ws['F2'] = f'=SUMIF(C1:C100,"={person2}",B1:B100)+SUMIF(C1:C100,"=Both",B1:B100)/2'

# Make spendings bold

bold_font = Font(bold=True)
ws['F1'].font = bold_font
ws['F2'].font = bold_font

# Add data validation

dv = DataValidation(type="list", formula1=f'"{person1},{person2},Both"', allow_blank=True)
ws.add_data_validation(dv)
dv.add('C1:C100')

# Add conditional colouring

red_font = Font(color='800000')
red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
green_font = Font(color='006400')
green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
yellow_font = Font(color='CCCC00')
yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

person1_rule = FormulaRule(formula=[f'C1="{person1}"'], fill=red_fill, font=red_font)
person2_rule = FormulaRule(formula=[f'C1="{person2}"'], fill=green_fill, font=green_font)
both_rule = FormulaRule(formula=['C1="Both"'], fill=yellow_fill, font=yellow_font)

ws.conditional_formatting.add('C1:C100', person1_rule)
ws.conditional_formatting.add('C1:C100', person2_rule)
ws.conditional_formatting.add('C1:C100', both_rule)

# Round numbers to 2 decimals

two_decimal_style = NamedStyle(name='two_decimals', number_format='0.00')

for number in range(1, 200):
    cell = 'B' + str(number)
    ws[cell].style = two_decimal_style

ws['F1'].style = two_decimal_style
ws['F2'].style = two_decimal_style

# Save Excel

current_time = datetime.now()
time = current_time.strftime("%H-%M-%S")
excel_name = "receipt-" + date + "-" + time + ".xlsx"
wb.save(excel_name)
